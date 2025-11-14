# app/blueprints/osversions/routes.py
from flask import render_template, request, jsonify, make_response
from . import osversions_bp
from velocitycmdb.app.utils.database import get_db_connection
import csv
from io import StringIO, BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


@osversions_bp.route('/')
def index():
    """OS Version Dashboard - Using existing device data only"""
    try:
        with get_db_connection() as conn:
            cursor = conn.cursor()

            # Get version distribution by vendor
            cursor.execute("""
                SELECT 
                    v.name as vendor,
                    d.os_version,
                    COUNT(d.id) as device_count,
                    GROUP_CONCAT(DISTINCT d.site_code) as sites,
                    MIN(d.id) as example_device_id,
                    MIN(d.name) as example_device_name
                FROM devices d
                LEFT JOIN vendors v ON d.vendor_id = v.id
                WHERE d.os_version IS NOT NULL 
                  AND d.os_version != ''
                  AND d.os_version != 'Unknown'
                GROUP BY v.name, d.os_version
                ORDER BY v.name, device_count DESC
            """)
            version_distribution = [dict(row) for row in cursor.fetchall()]

            # Get vendor summary with version counts
            cursor.execute("""
                SELECT 
                    v.name as vendor,
                    COUNT(DISTINCT d.id) as total_devices,
                    COUNT(DISTINCT d.os_version) as unique_versions
                FROM devices d
                LEFT JOIN vendors v ON d.vendor_id = v.id
                WHERE d.os_version IS NOT NULL 
                  AND d.os_version != ''
                  AND d.os_version != 'Unknown'
                GROUP BY v.name
                ORDER BY total_devices DESC
            """)
            vendor_summary = [dict(row) for row in cursor.fetchall()]

            # Get devices with missing/unknown versions
            cursor.execute("""
                SELECT 
                    d.id,
                    d.name,
                    d.site_code,
                    v.name as vendor,
                    d.model,
                    d.os_version
                FROM devices d
                LEFT JOIN vendors v ON d.vendor_id = v.id
                WHERE d.os_version IS NULL 
                   OR d.os_version = ''
                   OR d.os_version = 'Unknown'
                ORDER BY v.name, d.name
                LIMIT 50
            """)
            missing_versions = [dict(row) for row in cursor.fetchall()]

            # Calculate statistics
            cursor.execute("""
                SELECT 
                    COUNT(DISTINCT id) as total_devices,
                    COUNT(DISTINCT CASE 
                        WHEN os_version IS NOT NULL 
                         AND os_version != '' 
                         AND os_version != 'Unknown' 
                        THEN id END) as devices_with_version,
                    COUNT(DISTINCT os_version) as unique_versions
                FROM devices
            """)
            stats = dict(cursor.fetchone())
            stats['missing_count'] = stats['total_devices'] - stats['devices_with_version']
            stats['coverage_pct'] = round((stats['devices_with_version'] / stats['total_devices'] * 100), 1) if stats[
                                                                                                                    'total_devices'] > 0 else 0

            # Get data for pie chart - top 10 versions by device count
            cursor.execute("""
                SELECT 
                    v.name || ' - ' || d.os_version as version_label,
                    COUNT(d.id) as device_count
                FROM devices d
                LEFT JOIN vendors v ON d.vendor_id = v.id
                WHERE d.os_version IS NOT NULL 
                  AND d.os_version != ''
                  AND d.os_version != 'Unknown'
                GROUP BY v.name, d.os_version
                ORDER BY device_count DESC
                LIMIT 10
            """)
            chart_data = [dict(row) for row in cursor.fetchall()]

            return render_template('osversions/index.html',
                                   version_distribution=version_distribution,
                                   vendor_summary=vendor_summary,
                                   missing_versions=missing_versions,
                                   stats=stats,
                                   chart_data=chart_data)

    except Exception as e:
        print(f"ERROR: {str(e)}")
        import traceback
        traceback.print_exc()
        return render_template('osversions/index.html',
                               error=str(e),
                               version_distribution=[],
                               vendor_summary=[],
                               missing_versions=[],
                               stats={},
                               chart_data=[])


@osversions_bp.route('/vendor/<vendor_name>')
def vendor_detail(vendor_name):
    """Detailed view for a specific vendor's OS versions"""
    try:
        with get_db_connection() as conn:
            cursor = conn.cursor()

            # Get all versions for this vendor with device lists
            cursor.execute("""
                SELECT 
                    d.os_version,
                    COUNT(d.id) as device_count,
                    GROUP_CONCAT(d.name) as device_names,
                    GROUP_CONCAT(DISTINCT d.site_code) as sites,
                    GROUP_CONCAT(DISTINCT d.model) as models
                FROM devices d
                LEFT JOIN vendors v ON d.vendor_id = v.id
                WHERE v.name = ?
                  AND d.os_version IS NOT NULL 
                  AND d.os_version != ''
                  AND d.os_version != 'Unknown'
                GROUP BY d.os_version
                ORDER BY device_count DESC
            """, (vendor_name,))

            versions = []
            for row in cursor.fetchall():
                version_dict = dict(row)
                # Split concatenated fields into lists
                if version_dict['device_names']:
                    version_dict['devices'] = version_dict['device_names'].split(',')
                if version_dict['sites']:
                    version_dict['sites'] = list(set(version_dict['sites'].split(',')))
                if version_dict['models']:
                    version_dict['models'] = list(set(version_dict['models'].split(',')))
                versions.append(version_dict)

            return render_template('osversions/vendor_detail.html',
                                   vendor=vendor_name,
                                   versions=versions)

    except Exception as e:
        return render_template('osversions/vendor_detail.html',
                               error=str(e),
                               vendor=vendor_name,
                               versions=[])


@osversions_bp.route('/export')
def export_csv():
    """Export OS version report to Excel with two tabs"""
    try:
        with get_db_connection() as conn:
            cursor = conn.cursor()

            # Create workbook
            wb = Workbook()

            # Remove default sheet
            wb.remove(wb.active)

            # ========== TAB 1: Version Summary ==========
            ws1 = wb.create_sheet("Version Summary")

            # Header styling
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")

            # Tab 1 headers
            headers1 = ['Vendor', 'OS Version', 'Device Count', 'Sites', 'Example Device']
            ws1.append(headers1)

            # Style header row
            for cell in ws1[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')

            # Get version summary data
            cursor.execute("""
                SELECT 
                    v.name as vendor,
                    d.os_version,
                    COUNT(d.id) as device_count,
                    GROUP_CONCAT(DISTINCT d.site_code) as sites,
                    MIN(d.name) as example_device_name
                FROM devices d
                LEFT JOIN vendors v ON d.vendor_id = v.id
                WHERE d.os_version IS NOT NULL 
                  AND d.os_version != ''
                  AND d.os_version != 'Unknown'
                GROUP BY v.name, d.os_version
                ORDER BY v.name, device_count DESC
            """)

            for row in cursor.fetchall():
                vendor = row[0] or 'Unknown'
                os_version = row[1]
                device_count = row[2]
                sites = len(row[3].split(',')) if row[3] else 0
                example_device = row[4]

                ws1.append([vendor, os_version, device_count, sites, example_device])

            # Auto-size columns for tab 1
            for column in ws1.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws1.column_dimensions[column_letter].width = adjusted_width

            # ========== TAB 2: Device Details ==========
            ws2 = wb.create_sheet("Device Details")

            # Tab 2 headers
            headers2 = ['Device', 'Site', 'Vendor', 'Model', 'OS Version', 'Management IP']
            ws2.append(headers2)

            # Style header row
            for cell in ws2[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')

            # Get device details data
            cursor.execute("""
                SELECT 
                    d.name as device_name,
                    d.site_code,
                    v.name as vendor,
                    d.model,
                    d.os_version,
                    d.management_ip
                FROM devices d
                LEFT JOIN vendors v ON d.vendor_id = v.id
                WHERE d.os_version IS NOT NULL 
                  AND d.os_version != ''
                  AND d.os_version != 'Unknown'
                ORDER BY v.name, d.os_version, d.name
            """)

            for row in cursor.fetchall():
                ws2.append([row[0], row[1], row[2], row[3], row[4], row[5]])

            # Auto-size columns for tab 2
            for column in ws2.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws2.column_dimensions[column_letter].width = adjusted_width

            # Save to BytesIO
            output = BytesIO()
            wb.save(output)
            output.seek(0)

            response = make_response(output.read())
            response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            response.headers['Content-Disposition'] = 'attachment; filename=os_versions_report.xlsx'

            return response

    except Exception as e:
        print(f"Export error: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500